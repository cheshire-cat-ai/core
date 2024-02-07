import pandas as pd
from openpyxl import load_workbook
import xlrd
from abc import ABC
from typing import Iterator
from langchain.docstore.document import Document
from langchain.document_loaders.base import BaseBlobParser
from langchain.document_loaders.blob_loaders.schema import Blob

class XLSXExcelParser(BaseBlobParser, ABC):
    def lazy_parse(self, blob: Blob) -> Iterator[Document]:

        with blob.as_bytes_io() as file:
            if blob.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                wb = load_workbook(file, data_only=True)
            content = {}
            for page in wb.sheetnames:
                df = pd.read_excel(file, sheet_name=page, engine='openpyxl')   
                sheet_data_list = df.to_dict(orient='records')
                content[page] = sheet_data_list

        content = str(content)
        yield Document(page_content=content, metadata={})

class XLSExcelParser(BaseBlobParser, ABC):
    def lazy_parse(self, blob: Blob) -> Iterator[Document]:

        with blob.as_bytes_io() as file:
            if blob.mimetype == "application/vnd.ms-excel":
                wb = pd.read_excel(file, sheet_name=None, engine='xlrd')
            content = {}
            for sheet_name, sheet_data in wb.items():
                sheet_data_list = sheet_data.to_dict(orient='records')    
                content[sheet_name] = sheet_data_list

        content = str(content)
        yield Document(page_content=content, metadata={})